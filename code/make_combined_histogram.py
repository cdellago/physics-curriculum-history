#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Combined "Topics by decade" histogram comparing the Bachelor and Master
Physics curricula (University of Vienna, Version 2026).

Overlays the two decade distributions (Bachelor in blue, Master in orange)
on one axis. Years are read from column 8 of the "Topics by Module" sheet;
only numeric years are used (excludes "—" and the textual Euclid entry).

Produces:
  - BA_MA_Themen_Histogramm.png   (German)
  - BA_MA_Topics_Histogram.png    (English)

Requires: openpyxl, numpy, matplotlib.
Run:  python3 make_combined_histogram.py
"""

from pathlib import Path
from openpyxl import load_workbook
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# Resolve paths relative to the repository layout (code/, tables/, figures/)
BASE = Path(__file__).resolve().parent.parent
TBL = BASE / "tables"
FIG = BASE / "figures"


def read_years(xlsx_path, sheet="Topics by Module", year_col=8):
    ws = load_workbook(xlsx_path, data_only=True)[sheet]
    return np.array([int(ws.cell(r, year_col).value)
                     for r in range(2, ws.max_row + 1)
                     if isinstance(ws.cell(r, year_col).value, (int, float))])


def make_combined(ba, ma, lang, outfile):
    de = lang == "de"

    lo = (min(ba.min(), ma.min()) // 10) * 10
    hi = (max(ba.max(), ma.max()) // 10) * 10 + 10
    bins = np.arange(lo, hi + 10, 10)
    xticks = np.arange((lo // 20) * 20, hi + 20, 20)
    centuries = np.arange((lo // 100) * 100, hi + 100, 100)

    if de:
        lbl_ba = f"Bachelor (n = {len(ba)})"
        lbl_ma = f"Master (n = {len(ma)})"
        xlabel = "Jahrzehnt der Entdeckung / Einführung"
        ylabel = "Anzahl der Themen"
        title = ("BA & MA Physics (Version 2026): Themen nach Jahrzehnt\n"
                 f"Bachelor: Median {int(np.median(ba))}  |  "
                 f"Master: Median {int(np.median(ma))}")
    else:
        lbl_ba = f"Bachelor (n = {len(ba)})"
        lbl_ma = f"Master (n = {len(ma)})"
        xlabel = "Decade of discovery / introduction"
        ylabel = "Number of topics"
        title = ("BA & MA Physics (Version 2026): Topics by decade\n"
                 f"Bachelor: median {int(np.median(ba))}  |  "
                 f"Master: median {int(np.median(ma))}")

    fig, ax = plt.subplots(figsize=(11, 7.8))
    ax.hist(ba, bins=bins, color="tab:blue", alpha=0.8,
            edgecolor="white", linewidth=0.5, label=lbl_ba, zorder=3)
    ax.hist(ma, bins=bins, color="tab:orange", alpha=1.0,
            edgecolor="white", linewidth=0.5, label=lbl_ma, zorder=4)

    ax.set_xlabel(xlabel, fontsize=20)
    ax.set_ylabel(ylabel, fontsize=20)
    ax.set_title(title, fontsize=18)
    ax.set_xticks(xticks)
    ax.tick_params(axis="x", rotation=45, labelsize=15.5)
    ax.tick_params(axis="y", labelsize=15.5)
    ax.grid(axis="y", color="gray", alpha=0.3, zorder=0)
    for _c in centuries:                     # alternate century background shading
        if (_c // 100) % 2 == 1:
            ax.axvspan(max(_c, lo), min(_c + 100, hi), color="0.95", zorder=0)
    for c in centuries:
        ax.axvline(c, color="gray", linewidth=1.1, linestyle="--", alpha=0.7, zorder=1)
    ax.set_xlim(lo, hi)
    ax.margins(x=0.01)
    # label the centuries on a secondary top axis
    _ct = [c + 50 for c in centuries if lo <= c + 50 <= hi]
    _lab = []
    for _t in _ct:
        _n = (_t - 50) // 100 + 1
        _suf = "th" if 10 <= _n % 100 <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(_n % 10, "th")
        _lab.append(f"{_n}{_suf} century")
    _sax = ax.secondary_xaxis("top")
    _sax.set_xticks(_ct)
    _sax.set_xticklabels(_lab, fontsize=12.5, color="0.35", fontstyle="italic")
    _sax.tick_params(length=0)
    ax.legend(fontsize=14, frameon=True)

    fig.tight_layout()
    fig.savefig(outfile, dpi=150)
    plt.close(fig)
    return outfile


if __name__ == "__main__":
    ba = read_years(str(TBL / "BA_Physik_Module_Themen.xlsx"))
    ma = read_years(str(TBL / "MA_Physik_Module_Themen.xlsx"))
    for lang, outfile in [("de", "BA_MA_Themen_Histogramm.png"),
                          ("en", "BA_MA_Topics_Histogram.png")]:
        make_combined(ba, ma, lang, str(FIG / outfile))
        print(f"{outfile:32s}  BA n={len(ba)}  MA n={len(ma)}")
