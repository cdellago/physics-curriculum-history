#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate "Topics by decade" histograms for the Vienna Physics curricula.

Produces four PNGs from the topic spreadsheets:
  - BA German : BA_Physik_Themen_Histogramm.png
  - BA English: BA_Physics_Topics_Histogram.png
  - MA German : MA_Physik_Themen_Histogramm.png
  - MA English: MA_Physics_Topics_Histogram.png

The year is read from column 8 ("Jahr" / "Year") of the "Topics by Module"
sheet. Only numeric year cells are used, which automatically excludes
non-attributable topics ("—") and the textual "≈ 300 v. Chr." entry (Euclid).

Requires: openpyxl, numpy, matplotlib.
Run:  python3 make_histograms.py
"""

from pathlib import Path
from openpyxl import load_workbook
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# Resolve paths relative to the repository layout (code/, tables/, figures/)
BASE = Path(__file__).resolve().parent.parent
TBL = BASE / "tables"
FIG = BASE / "figures"


def read_years(xlsx_path, sheet="Topics by Module", year_col=8):
    """Return a numpy array of the numeric years in the given sheet/column."""
    ws = load_workbook(xlsx_path, data_only=True)[sheet]
    years = [int(ws.cell(r, year_col).value)
             for r in range(2, ws.max_row + 1)
             if isinstance(ws.cell(r, year_col).value, (int, float))]
    return np.array(years)


def make_histogram(years, lang, level, note, outfile):
    """
    Draw one decade histogram.

    years   : array of integer years
    lang    : "de" or "en"
    level   : "BA" or "MA"
    note    : extra string for the subtitle (e.g. "ohne Euklid" / "excl. Euclid"),
              or "" for none
    outfile : output PNG path
    """
    de = lang == "de"

    # decade bins; labelled ticks on multiples of 20; gridlines one century apart
    lo = (years.min() // 10) * 10
    hi = (years.max() // 10) * 10 + 10
    bins = np.arange(lo, hi + 10, 10)
    xticks = np.arange((years.min() // 20) * 20, hi + 20, 20)
    centuries = np.arange((years.min() // 100) * 100, hi + 100, 100)

    fig, ax = plt.subplots(figsize=(9.5, 7.8))
    ax.hist(years, bins=bins, color="tab:blue",
            edgecolor="white", linewidth=0.6, zorder=3)

    median = int(np.median(years))
    sep = f" ({note})" if note else ""
    if de:
        ax.set_xlabel("Jahrzehnt der Entdeckung / Einführung", fontsize=20)
        ax.set_ylabel("Anzahl der Themen", fontsize=20)
        ax.set_title(f"{level} Physik (Version 2026): Themen nach Jahrzehnt\n"
                     f"n = {len(years)} datierte Themen{sep}; Median {median}",
                     fontsize=19)
    else:
        ax.set_xlabel("Decade of discovery / introduction", fontsize=20)
        ax.set_ylabel("Number of topics", fontsize=20)
        ax.set_title(f"{level} Physics (Version 2026): Topics by decade\n"
                     f"n = {len(years)} dated topics{sep}; median {median}",
                     fontsize=19)

    ax.set_xticks(xticks)
    ax.tick_params(axis="x", rotation=45, labelsize=15.5)
    ax.tick_params(axis="y", labelsize=15.5)
    ax.grid(axis="y", color="gray", alpha=0.3, zorder=0)
    for c in centuries:                      # vertical gridlines, one century apart
        ax.axvline(c, color="gray", linewidth=1.1, linestyle="--", alpha=0.7, zorder=1)
    ax.set_xlim(lo, hi)
    ax.margins(x=0.01)
    # label the centuries on a secondary top axis
    _ct = [c + 50 for c in centuries if lo <= c + 50 <= hi]
    _lab = []
    for _t in _ct:
        _n = (_t - 50) // 100 + 1
        _suf = "th" if 10 <= _n % 100 <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(_n % 10, "th")
        _lab.append(f"{_n}{_suf} century")
    _sax = ax.secondary_xaxis("top")
    _sax.set_xticks(_ct)
    _sax.set_xticklabels(_lab, fontsize=12.5, color="0.35", fontstyle="italic")
    _sax.tick_params(length=0)

    fig.tight_layout()
    fig.savefig(outfile, dpi=150)
    plt.close(fig)
    return outfile


JOBS = [
    # (xlsx, level, lang, note, outfile)
    ("BA_Physik_Module_Themen.xlsx", "BA", "de", "ohne Euklid",  "BA_Physik_Themen_Histogramm.png"),
    ("BA_Physik_Module_Themen.xlsx", "BA", "en", "excl. Euclid", "BA_Physics_Topics_Histogram.png"),
    ("MA_Physik_Module_Themen.xlsx", "MA", "de", "",             "MA_Physik_Themen_Histogramm.png"),
    ("MA_Physik_Module_Themen.xlsx", "MA", "en", "",             "MA_Physics_Topics_Histogram.png"),
]

if __name__ == "__main__":
    for xlsx, level, lang, note, outfile in JOBS:
        years = read_years(str(TBL / xlsx))
        make_histogram(years, lang, level, note, str(FIG / outfile))
        print(f"{outfile:40s}  n={len(years):3d}  "
              f"median={int(np.median(years))}  mean={years.mean():.0f}")
